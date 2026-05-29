"""生成 ITSM 变更工单 Excel（四 Sheet）。"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="E8E8E8")
VENDOR_FILLS = {
    "华为": PatternFill("solid", fgColor="DCFCE7"),
    "H3C": PatternFill("solid", fgColor="DBEAFE"),
    "H3CF1000": PatternFill("solid", fgColor="DBEAFE"),
}


def _style_header_row(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def build_change_ticket_workbook(
    manifest: dict[str, Any],
    *,
    config_zip_url: str | None = None,
    workflow_run_id: str | None = None,
) -> bytes:
    wb = Workbook()

    # Sheet1 主表
    ws1 = wb.active
    ws1.title = "变更工单"
    main_rows = [
        ("变更编号", manifest.get("ticket_id", "")),
        ("变更标题", manifest.get("ticket_title", "")),
        ("变更背景", manifest.get("change_background", "")),
        ("变更目的", manifest.get("change_purpose", "")),
        ("变更申请人", manifest.get("requester", "")),
        ("变更申请部门", manifest.get("requester_dept", "")),
        ("变更优先级", manifest.get("priority", "P2")),
        ("计划执行时间", manifest.get("due_date") or ""),
        ("变更设备数量", manifest.get("device_count", len(manifest.get("devices", [])))),
        ("变更类型", manifest.get("change_type", "配置变更")),
        ("变更风险等级", manifest.get("risk_level", "中")),
        ("策略 ZIP 附件", config_zip_url or ""),
        ("Workflow Run ID", workflow_run_id or ""),
    ]
    ws1.append(["字段", "值"])
    for row in main_rows:
        ws1.append(list(row))
    _style_header_row(ws1, 2)
    _autosize(ws1, 60)

    # Sheet2 设备
    ws2 = wb.create_sheet("变更设备列表")
    dev_headers = [
        "设备名称",
        "IP地址",
        "厂商",
        "型号",
        "当前版本",
        "变更前配置摘要",
        "变更后配置摘要",
    ]
    ws2.append(dev_headers)
    for d in manifest.get("devices", []):
        ws2.append(
            [
                d.get("device_name", ""),
                d.get("ip_address", ""),
                d.get("vendor", ""),
                d.get("model", ""),
                d.get("version", ""),
                d.get("before_summary", ""),
                d.get("after_summary", ""),
            ]
        )
    _style_header_row(ws2, len(dev_headers))
    _autosize(ws2)

    # Sheet3 脚本
    ws3 = wb.create_sheet("变更脚本")
    script_headers = ["设备名称", "厂商", "执行顺序", "配置命令", "命令数量"]
    ws3.append(script_headers)
    for s in manifest.get("scripts", []):
        vendor = s.get("vendor", "")
        row_idx = ws3.max_row + 1
        ws3.append(
            [
                s.get("device_name", ""),
                vendor,
                s.get("order", ""),
                s.get("commands", ""),
                s.get("command_count", 0),
            ]
        )
        fill = VENDOR_FILLS.get(vendor)
        if fill:
            for col in range(1, len(script_headers) + 1):
                ws3.cell(row=row_idx, column=col).fill = fill
        ws3.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    _style_header_row(ws3, len(script_headers))
    ws3.column_dimensions["D"].width = 72

    # Sheet4 回退
    ws4 = wb.create_sheet("回退方案")
    rb_headers = ["步骤", "设备", "回退命令", "预期效果", "执行人"]
    ws4.append(rb_headers)
    for r in manifest.get("rollback", []):
        ws4.append(
            [
                r.get("step", ""),
                r.get("device_name", ""),
                r.get("rollback_command", ""),
                r.get("expected_effect", ""),
                r.get("executor", ""),
            ]
        )
    _style_header_row(ws4, len(rb_headers))
    _autosize(ws4)

    footer = (
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
        f"trace_id: {manifest.get('trace_id') or workflow_run_id or '-'}"
    )
    ws1.append([])
    ws1.append(["备注", footer])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
