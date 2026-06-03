# SPDX-FileCopyrightText: 2026 wangdong <wangdong5919@163.com>
# SPDX-License-Identifier: Apache-2.0

"""基于金融行业标准模板生成 ITSM 变更工单 Excel（Skill 内实现）。"""

from __future__ import annotations

import io
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

SKILL_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = SKILL_ROOT / "templates" / "change_ticket_template.xlsx"

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="E8E8E8")
VENDOR_FILLS = {
    "华为": PatternFill("solid", fgColor="DCFCE7"),
    "H3C": PatternFill("solid", fgColor="DBEAFE"),
    "H3CF1000": PatternFill("solid", fgColor="DBEAFE"),
}


def _copy_cell_style(src: Cell, dst: Cell) -> None:
    if not src.has_style:
        return
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def _snapshot_row(ws: Worksheet, row_idx: int, col_count: int = 7) -> list[Cell]:
    return [ws.cell(row_idx, col) for col in range(1, col_count + 1)]


def _write_row(
    ws: Worksheet,
    row_idx: int,
    values: list[Any],
    style_cells: list[Cell] | None = None,
) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        if style_cells and col <= len(style_cells):
            _copy_cell_style(style_cells[col - 1], cell)
    if style_cells and len(style_cells) >= 6:
        cmd_cell = ws.cell(row=row_idx, column=6)
        if cmd_cell.value and isinstance(cmd_cell.value, str) and "\n" in cmd_cell.value:
            cmd_cell.alignment = Alignment(wrap_text=True, vertical="top")


def _parse_due_date(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return text


def _build_verification_steps(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    steps: list[dict[str, Any]] = []
    seq = 0
    verifier = manifest.get("assignee") or manifest.get("requester") or "运维"
    for script in manifest.get("scripts") or []:
        seq += 1
        device = script.get("device_name", "")
        vendor = str(script.get("vendor") or "")
        if vendor in ("华为", "Huawei"):
            method = "display acl all"
        elif vendor in ("H3C", "H3CF1000"):
            method = "display current-configuration | include object-group"
        else:
            method = f"show running-config | include {device}"
        steps.append(
            {
                "step": seq,
                "title": f"验证 {device} 策略是否生效",
                "method": method,
                "expected": "配置与变更脚本一致，策略命中正常",
                "verifier": verifier,
                "standard": "规则条目/对象组与工单附件一致",
            }
        )
    seq += 1
    steps.append(
        {
            "step": seq,
            "title": "业务连通性验证",
            "method": "ping / telnet / 业务探测",
            "expected": "业务流量正常通过，无丢包或拒绝",
            "verifier": manifest.get("requester") or "业务方",
            "standard": "符合变更目的中的连通性要求",
        }
    )
    return steps


def _fill_static_header(ws: Worksheet, manifest: dict[str, Any]) -> None:
    device_count = manifest.get("device_count") or len(manifest.get("devices") or [])
    ws.cell(row=2, column=2, value=manifest.get("ticket_id", ""))
    ws.cell(row=2, column=4, value=manifest.get("ticket_title", ""))
    ws.cell(row=2, column=6, value=manifest.get("change_type", "配置变更"))
    ws.cell(row=2, column=7, value=manifest.get("risk_level", "中"))
    ws.cell(row=3, column=2, value=manifest.get("change_background", ""))
    ws.cell(row=4, column=2, value=manifest.get("change_purpose", ""))
    ws.cell(row=5, column=2, value=manifest.get("requester", ""))
    ws.cell(row=5, column=4, value=manifest.get("requester_dept", ""))
    ws.cell(row=5, column=6, value=_parse_due_date(manifest.get("due_date")))
    ws.cell(row=5, column=7, value=manifest.get("assignee") or manifest.get("requester") or "")
    ws.cell(row=6, column=2, value=manifest.get("technical_reviewer") or "-")
    ws.cell(row=6, column=4, value=manifest.get("reviewer") or "-")
    ws.cell(row=6, column=6, value=device_count)


def _append_section_header(ws: Worksheet, row_idx: int, header_cells: list[Cell]) -> int:
    values = [c.value for c in header_cells]
    _write_row(ws, row_idx, values, header_cells)
    return row_idx + 1


def _find_section_row(ws: Worksheet, label: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == label:
            return row
    raise ValueError(f"模板缺少区块: {label}")


def _build_from_template(manifest: dict[str, Any]) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ws.title = "变更工单"

    _fill_static_header(ws, manifest)

    device_hdr_row = _find_section_row(ws, "变更设备清单")
    script_hdr_row = _find_section_row(ws, "变更执行脚本")
    verify_hdr_row = _find_section_row(ws, "变更验证环节")
    rollback_hdr_row = _find_section_row(ws, "回退方案")

    device_data_style = _snapshot_row(ws, device_hdr_row + 1)
    script_header = _snapshot_row(ws, script_hdr_row)
    script_data_style = _snapshot_row(ws, script_hdr_row + 1)
    verify_header = _snapshot_row(ws, verify_hdr_row)
    verify_data_style = _snapshot_row(ws, verify_hdr_row + 1)
    rollback_header = _snapshot_row(ws, rollback_hdr_row)
    rollback_data_style = _snapshot_row(ws, rollback_hdr_row + 1)

    if ws.cell(1, 1).value and str(ws.cell(1, 1).value).startswith("列"):
        ws.delete_rows(1, 1)
        device_hdr_row = _find_section_row(ws, "变更设备清单")

    if ws.max_row > device_hdr_row:
        ws.delete_rows(device_hdr_row + 1, ws.max_row - device_hdr_row)

    row = device_hdr_row + 1
    devices = manifest.get("devices") or []
    for device in devices:
        _write_row(
            ws,
            row,
            [
                "-",
                device.get("device_name", ""),
                device.get("ip_address", ""),
                device.get("vendor", ""),
                device.get("model", ""),
                "-",
                "-",
            ],
            device_data_style,
        )
        row += 1

    row = _append_section_header(ws, row, script_header)
    scripts = manifest.get("scripts") or []
    for idx, script in enumerate(scripts, start=1):
        _write_row(
            ws,
            row,
            [
                "-",
                idx,
                script.get("device_name", ""),
                script.get("vendor", ""),
                script.get("order", idx),
                script.get("commands", ""),
                "-",
            ],
            script_data_style,
        )
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, 60)
        row += 1

    row = _append_section_header(ws, row, verify_header)
    for step in _build_verification_steps(manifest):
        _write_row(
            ws,
            row,
            [
                "-",
                step["step"],
                step["title"],
                step["method"],
                step["expected"],
                step["verifier"],
                step["standard"],
            ],
            verify_data_style,
        )
        row += 1

    row = _append_section_header(ws, row, rollback_header)
    rollback_rows = list(manifest.get("rollback") or [])
    if not rollback_rows and scripts:
        for idx, script in enumerate(scripts, start=1):
            rollback_rows.append(
                {
                    "step": idx,
                    "device_name": script.get("device_name", ""),
                    "rollback_command": "rollback configuration",
                    "expected_effect": "回滚到变更前配置",
                    "duration": "5 分钟",
                    "executor": manifest.get("assignee") or manifest.get("requester") or "运维",
                }
            )
    for rb in rollback_rows:
        _write_row(
            ws,
            row,
            [
                "-",
                rb.get("step", ""),
                rb.get("device_name", ""),
                rb.get("rollback_command", ""),
                rb.get("expected_effect", ""),
                rb.get("duration") or rb.get("estimated_time") or "5 分钟",
                rb.get("executor", ""),
            ],
            rollback_data_style,
        )
        row += 1

    footer_row = row + 1
    trace = manifest.get("trace_id") or manifest.get("workflow_run_id") or "-"
    ws.cell(
        row=footer_row,
        column=1,
        value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | trace_id: {trace}",
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_change_ticket_workbook(
    manifest: dict[str, Any],
    *,
    workflow_run_id: str | None = None,
) -> bytes:
    data = dict(manifest)
    if workflow_run_id:
        data.setdefault("workflow_run_id", workflow_run_id)
    if TEMPLATE_PATH.is_file():
        return _build_from_template(data)
    return _build_legacy_workbook(data, workflow_run_id=workflow_run_id)


def _style_header_row(ws: Worksheet, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _build_legacy_workbook(manifest: dict[str, Any], *, workflow_run_id: str | None = None) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "变更工单"
    ws1.append(["字段", "值"])
    for row in [
        ("变更编号", manifest.get("ticket_id", "")),
        ("变更标题", manifest.get("ticket_title", "")),
        ("变更背景", manifest.get("change_background", "")),
        ("变更目的", manifest.get("change_purpose", "")),
    ]:
        ws1.append(list(row))
    _style_header_row(ws1, 2)
    _autosize(ws1, 60)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
