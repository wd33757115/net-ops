#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
patrol_tool_opt.py
网络设备巡检工具 - 严格使用数据库独立用户名和密码
联系邮箱: wangdong5919@163.com
Copyright [2025] [wangdong]
SPDX-License-Identifier: Apache-2.0
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sqlite3
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone, timedelta
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from src.core.patrol.textfsm_assets import discover_textfsm_templates

# ---------- 导入设备管理模块 ----------
from device_manager import DBManager as DeviceDBManager

# ---------- Netmiko ----------
try:
    from netmiko import ConnectHandler
except ImportError:
    raise ImportError("Netmiko 模块未安装，请安装以支持设备连接")

# ---------- TextFSM ----------
try:
    import textfsm
except ImportError:
    textfsm = None
    logging.warning("TextFSM 模块不可用，结构化解析将被禁用")

# ---------- YAML ----------
try:
    import yaml
except ImportError:
    yaml = None
    logging.warning("YAML 模块不可用，将使用 JSON 配置回退")

# -----------------------
# 常量 / 默认值
# -----------------------
DEFAULT_DB = Path("db/patrol.db")
DEVICES_DB = Path("db/devices.db")
LOG_DIR = Path("logs")
BASELINE_TABLE = "baseline_data"
PATROL_TABLE = "patrol_data"
DIFF_TABLE = "diff_data"
STATS_TABLE = "patrol_stats"

# Excel 样式
HEADER_FILL = PatternFill(start_color="BDD7EE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

# 时区
BEIJING_TZ = timezone(timedelta(hours=8))

# -----------------------
# 状态枚举
# -----------------------
class State(Enum):
    UNCHANGED = "不变"
    MODIFIED = "修改"
    ADDED = "新增"
    DELETED = "缺失"
    FAILED = "失败"

# 状态优先级
STATE_PRIORITY = {
    State.MODIFIED.value: 3,
    State.ADDED.value: 2,
    State.DELETED.value: 2,
    State.UNCHANGED.value: 1,
    State.FAILED.value: 0,
}

# -----------------------
# 日志配置
# -----------------------
LOG_DIR.mkdir(exist_ok=True)

# -----------------------
# 工具函数
# -----------------------
def load_config(path: Optional[Path]) -> Dict[str, Any]:
    if not path or not path.exists():
        return {}
    try:
        if path.suffix.lower() in (".yml", ".yaml") and yaml:
            return yaml.safe_load(path.read_text(encoding='utf-8'))
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception as e:
        logger.error(f"加载配置失败: {e}")
        return {}

def safe_sheet_name(name: str) -> str:
    import re
    s = re.sub(r'[\[\]:*?/\\]', "_", str(name).strip())
    return s[:31]

def cmd_to_template_name(cmd: str) -> str:
    return cmd.strip().lower().replace(" ", "_") + ".textfsm"

def sanitize_for_excel(text: Optional[str]) -> str:
    """清理字符串，使其可安全写入 Excel"""
    if text is None:
        return ""
    text = ''.join(c for c in text if c >= ' ' or c in {'\n', '\t'})
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    text = text.replace('"', '&quot;').replace("'", '&apos;')
    return text[:32767]

# -----------------------
# 巡检结果 DBManager
# -----------------------
class PatrolDBManager:
    def __init__(self, db_file: Path = DEFAULT_DB):
        self.db_file = db_file
        self.conn = sqlite3.connect(str(self.db_file), timeout=30, check_same_thread=False)
        self._init_db()

    def _init_db(self):
        cur = self.conn.cursor()
        for table, sql in [
            (BASELINE_TABLE, """
                CREATE TABLE IF NOT EXISTS baseline_data (
                    device_id TEXT NOT NULL,
                    command TEXT NOT NULL,
                    structured TEXT,
                    text_output TEXT,
                    last_update TEXT,
                    PRIMARY KEY (device_id, command)
                )
            """),
            (PATROL_TABLE, """
                CREATE TABLE IF NOT EXISTS patrol_data (
                    run_id TEXT NOT NULL,
                    device_id TEXT NOT NULL,
                    command TEXT NOT NULL,
                    structured TEXT,
                    text_output TEXT,
                    timestamp TEXT,
                    PRIMARY KEY (run_id, device_id, command)
                )
            """),
            (DIFF_TABLE, """
                CREATE TABLE IF NOT EXISTS diff_data (
                    run_id TEXT NOT NULL,
                    device_id TEXT NOT NULL,
                    command TEXT NOT NULL,
                    state TEXT,
                    structured_diff TEXT,
                    text_diff TEXT,
                    text_output TEXT,
                    baseline_text TEXT,
                    timestamp TEXT,
                    PRIMARY KEY (run_id, device_id, command)
                )
            """),
            (STATS_TABLE, """
                CREATE TABLE IF NOT EXISTS patrol_stats (
                    run_id TEXT,
                    device_id TEXT,
                    total_commands INTEGER,
                    added INTEGER,
                    deleted INTEGER,
                    modified INTEGER,
                    timestamp TEXT,
                    PRIMARY KEY (run_id, device_id)
                )
            """),
        ]:
            cur.execute(sql)
        for idx in [
            f"idx_baseline_device_id ON {BASELINE_TABLE}(device_id)",
            f"idx_patrol_run_id ON {PATROL_TABLE}(run_id)",
            f"idx_diff_run_id ON {DIFF_TABLE}(run_id)",
            f"idx_stats_run_id ON {STATS_TABLE}(run_id)",
        ]:
            cur.execute(f"CREATE INDEX IF NOT EXISTS {idx}")
        self.conn.commit()

    def save_baseline_results(self, rows: List[Tuple[str, str, Optional[str], Optional[str]]]):
        if not rows: return
        cur = self.conn.cursor()
        now = datetime.now(BEIJING_TZ).isoformat()
        cur.execute("BEGIN")
        for dev_id, cmd, structured, text in rows:
            cur.execute(f"""
            INSERT OR REPLACE INTO {BASELINE_TABLE}
            (device_id, command, structured, text_output, last_update)
            VALUES (?,?,?,?,?)
            """, (dev_id, cmd, structured, text, now))
        cur.execute("COMMIT")

    def save_patrol_results(self, run_id: str, rows: List[Tuple[str, str, Optional[str], Optional[str]]]):
        if not rows: return
        cur = self.conn.cursor()
        ts = datetime.now(BEIJING_TZ).isoformat()
        cur.execute("BEGIN")
        for dev_id, cmd, structured, text in rows:
            cur.execute(f"""
            INSERT OR REPLACE INTO {PATROL_TABLE}
            (run_id, device_id, command, structured, text_output, timestamp)
            VALUES (?,?,?,?,?,?)
            """, (run_id, dev_id, cmd, structured, text, ts))
        cur.execute("COMMIT")

    def save_diff_results(self, run_id: str, rows: List[Tuple[str, str, str, Optional[str], Optional[str], Optional[str], Optional[str]]]):
        if not rows: return
        cur = self.conn.cursor()
        ts = datetime.now(BEIJING_TZ).isoformat()
        cur.execute("BEGIN")
        for dev_id, cmd, state, sdiff, tdiff, text, base in rows:
            cur.execute(f"""
            INSERT OR REPLACE INTO {DIFF_TABLE}
            (run_id, device_id, command, state, structured_diff, text_diff, text_output, baseline_text, timestamp)
            VALUES (?,?,?,?,?,?,?,?,?)
            """, (run_id, dev_id, cmd, state, sdiff, tdiff, text, base, ts))
        cur.execute("COMMIT")

    def save_stats(self, run_id: str, dev_id: str, total: int, added: int, deleted: int, modified: int):
        cur = self.conn.cursor()
        ts = datetime.now(BEIJING_TZ).isoformat()
        cur.execute(f"""
        INSERT OR REPLACE INTO {STATS_TABLE}
        (run_id, device_id, total_commands, added, deleted, modified, timestamp)
        VALUES (?,?,?,?,?,?,?)
        """, (run_id, dev_id, total, added, deleted, modified, ts))
        self.conn.commit()

    def get_baseline_entry(self, dev_id: str, cmd: str) -> Tuple[Optional[str], Optional[str]]:
        cur = self.conn.cursor()
        cur.execute(f"SELECT structured, text_output FROM {BASELINE_TABLE} WHERE device_id=? AND command=?", (dev_id, cmd))
        row = cur.fetchone()
        return (row[0], row[1]) if row else (None, None)

    def load_diff_for_report(self, run_id: Optional[str] = None) -> Dict[str, List[Dict]]:
        cur = self.conn.cursor()
        sql = f"SELECT run_id, device_id, command, state, structured_diff, text_diff, text_output, baseline_text FROM {DIFF_TABLE}"
        params = []
        if run_id:
            sql += " WHERE run_id=?"
            params = [run_id]
        cur.execute(sql, params)
        data = {}
        for row in cur.fetchall():
            sheet = safe_sheet_name(row[2])
            data.setdefault(sheet, []).append({
                "run_id": row[0],
                "device_id": row[1],
                "command": row[2],
                "state": row[3],
                "structured_diff": json.loads(row[4]) if row[4] else None,
                "text_diff": json.loads(row[5]) if row[5] else None,
                "text_output": row[6],
                "baseline_text": row[7],
            })
        return data

    def get_command_stats(self, run_id: Optional[str] = None) -> List[Dict]:
        cur = self.conn.cursor()
        sql = f"SELECT run_id, device_id, command, state, structured_diff, text_diff FROM {DIFF_TABLE}"
        params = []
        if run_id:
            sql += " WHERE run_id=?"
            params = [run_id]
        cur.execute(sql, params)
        stats = []
        for row in cur.fetchall():
            a = d = m = 0
            for diff in (row[4], row[5]):
                if not diff: continue
                for r in json.loads(diff):
                    st = r.get("状态", State.UNCHANGED.value)
                    if st == State.ADDED.value: a += 1
                    elif st == State.DELETED.value: d += 1
                    elif st == State.MODIFIED.value: m += 1
            stats.append({
                "run_id": row[0],
                "device_id": row[1],
                "command": row[2],
                "added": a,
                "deleted": d,
                "modified": m,
            })
        return stats

    def get_failed_devices(self, run_id: Optional[str] = None) -> List[str]:
        cur = self.conn.cursor()
        sql = f"SELECT device_id FROM {STATS_TABLE} WHERE total_commands>0 AND added=0 AND deleted=0 AND modified=0"
        params = []
        if run_id:
            sql += " AND run_id=?"
            params = [run_id]
        cur.execute(sql, params)
        return [r[0] for r in cur.fetchall()]

    def close(self):
        self.conn.close()

# -----------------------
# 差异算法
# -----------------------
def parse_with_textfsm(fsm_path: Path, text: str) -> List[Dict]:
    if not textfsm:
        raise RuntimeError("textfsm 未安装")
    with fsm_path.open("r", encoding="utf-8") as f:
        fsm = textfsm.TextFSM(f)
        parsed = fsm.ParseText(text or "")
    headers = [h.upper() for h in fsm.header]
    return [dict(zip(headers, row)) for row in parsed]

def diff_text_lines(base: Optional[str], curr: Optional[str]) -> List[Dict]:
    import difflib
    b = (base or "").splitlines()
    c = (curr or "").splitlines()
    sm = difflib.SequenceMatcher(None, b, c)
    out = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            for j in range(j1, j2):
                out.append({"内容": c[j], "状态": State.UNCHANGED.value})
        elif tag == "replace":
            for j in range(j1, j2):
                out.append({"内容": c[j], "状态": State.MODIFIED.value})
            if i2 - i1 > j2 - j1:
                for i in range(j1 + (j2 - j1), i2):
                    out.append({"内容": b[i], "状态": State.DELETED.value})
        elif tag == "delete":
            for i in range(i1, i2):
                out.append({"内容": b[i], "状态": State.DELETED.value})
        elif tag == "insert":
            for j in range(j1, j2):
                out.append({"内容": c[j], "状态": State.ADDED.value})
    return out

def diff_structured_lcs(base_rows: List[Dict], curr_rows: List[Dict], key_cols: List[str]) -> List[Dict]:
    if not key_cols:
        possible = ['ID', 'NAME', 'INDEX', 'PORT', 'INTERFACE']
        key_cols = [k for k in possible if k in (base_rows[0] if base_rows else {}) and k in (curr_rows[0] if curr_rows else {})]
        if not key_cols and base_rows:
            key_cols = list(base_rows[0].keys())[:2]

    def sort_key(r): return tuple(str(r.get(k, "")) for k in key_cols)
    base_rows = sorted(base_rows, key=sort_key)
    curr_rows = sorted(curr_rows, key=sort_key)

    def key_of(r): return "|".join(str(r.get(k, "")) for k in key_cols)
    def rows_equal(a, b):
        excl = {"状态"}
        return {k: str(v).strip() for k, v in a.items() if k not in excl} == \
               {k: str(v).strip() for k, v in b.items() if k not in excl}

    bk = [key_of(r) for r in base_rows]
    ck = [key_of(r) for r in curr_rows]

    m, n = len(bk), len(ck)
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if bk[i - 1] == ck[j - 1]:
                dp[i][j] = dp[i - 1][j - 1] + 1
            else:
                dp[i][j] = max(dp[i - 1][j], dp[i][j - 1])

    i, j = m, n
    res = []
    while i > 0 or j > 0:
        if i > 0 and j > 0 and bk[i - 1] == ck[j - 1]:
            out = dict(curr_rows[j - 1])
            out["状态"] = State.UNCHANGED.value if rows_equal(base_rows[i - 1], curr_rows[j - 1]) else State.MODIFIED.value
            res.append(out)
            i -= 1; j -= 1
        elif j > 0 and (i == 0 or dp[i][j - 1] >= dp[i - 1][j]):
            out = dict(curr_rows[j - 1])
            out["状态"] = State.ADDED.value
            res.append(out)
            j -= 1
        else:
            out = dict(base_rows[i - 1])
            out["状态"] = State.DELETED.value
            res.append(out)
            i -= 1
    return list(reversed(res))

# -----------------------
# 设备巡检核心
# -----------------------
class DevicePatroller:
    def __init__(self, patrol_db: PatrolDBManager, device_db: DeviceDBManager,
                 templates_dir: Path, commands_dir: Path,
                 config: Dict, force_baseline: bool):
        self.patrol_db = patrol_db
        self.device_db = device_db
        self.templates_dir = templates_dir
        self.commands_dir = commands_dir
        self.config = config
        self.force_baseline = force_baseline
        self.max_retries = int(config.get("max_retries", 1))
        self.retry_delay = float(config.get("retry_delay", 1.0))
        self.structured_cmds = config.get("structured_commands", [])
        self.key_map = config.get("primary_keys", {})
        self.templates_cache = {}

    def _discover_templates(self, model: str) -> Dict[str, Path]:
        if model in self.templates_cache:
            return self.templates_cache[model]
        mapping = discover_textfsm_templates(
            model=model,
            legacy_root=self.templates_dir,
        )
        self.templates_cache[model] = mapping
        return mapping

    def _strip_command_and_prompt(self, raw: str, cmd: str, dev_id: str) -> str:
        lines = raw.splitlines()
        cleaned = []
        cmd_norm = cmd.strip()
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if line == cmd_norm or line.startswith(cmd_norm):
                i += 1
                continue
            if re.match(r'^[-*_\s]+$', line):
                i += 1
                continue
            break

        prompt_pat = re.compile(r'^[<[][^]>]+[>]\s*$')
        for line in lines[i:]:
            stripped = line.rstrip()
            if prompt_pat.match(stripped):
                break
            cleaned.append(stripped)

        result = '\n'.join(cleaned).strip()
        logger.debug(f"{dev_id} 输出清理: {len(raw)} → {len(result)}")
        return result

    def _ssh_execute(self, conn, cmd: str, dev_id: str) -> Tuple[bool, str]:
        device_type = conn.device_type.lower()
        expect_config = self.config.get("expect_patterns", {})
        expect_str = expect_config.get("default", r"[>#\$%]")

        for key, pattern in expect_config.items():
            if key != "default" and key in device_type:
                expect_str = pattern
                break

        logger.debug(f"{dev_id} 使用 expect_string: {expect_str}")

        try:
            output = conn.send_command(
                cmd,
                expect_string=expect_str,
                read_timeout=400,
                strip_prompt=False,
                strip_command=False,
            )
            logger.info(f"{dev_id} expect_string 模式成功: {cmd}")
            clean = self._strip_command_and_prompt(output, cmd, dev_id)
            return True, clean
        except Exception as e:
            logger.warning(f"{dev_id} expect_string 失败: {e}，尝试 timing 模式")
            try:
                output = conn.send_command_timing(
                    cmd,
                    read_timeout=400,
                    delay_factor_compat=True,
                    delay_factor=8.0,
                    strip_prompt=False,
                    strip_command=False,
                )
                logger.info(f"{dev_id} fallback timing 模式成功: {cmd}")
                clean = self._strip_command_and_prompt(output, cmd, dev_id)
                return True, clean
            except Exception as e2:
                logger.error(f"{dev_id} 两种执行方式均失败: {cmd} - {e2}")
                return False, f"命令执行失败（两种方式均失败）\n主错误：{e}\nFallback错误：{e2}"

    def _create_failed_result(self, dev_id: str, reason: str, total: int = 0, cmds: List[str] = None):
        """统一失败结果处理"""
        if cmds is None:
            cmds = []
        run_id = self.config.get("run_id")
        diff_batch = []
        for cmd in cmds:
            diff_batch.append((dev_id, cmd, State.FAILED.value, None,
                               json.dumps([{"内容": reason, "状态": State.FAILED.value}]), None, None))
        self.patrol_db.save_diff_results(run_id, diff_batch)
        self.patrol_db.save_stats(run_id, dev_id, total, 0, 0, 0)
        return {"device_id": dev_id, "total": total, "added": 0, "deleted": 0, "modified": 0}

    def patrol_single_device(self, dev: Dict) -> Dict:
        start = time.time()
        dev_name = dev.get("device_name", "unknown")
        ip = dev.get("ip", "unknown")
        dev_id = f"{dev_name}-{ip}"

        # ==================== 严格使用数据库独立用户名和密码 ====================
        device_username = dev.get("username")
        device_password = dev.get("password")

        if not device_username:
            logger.error(f"{dev_id} 缺少用户名！（数据库 username 字段为空）")
            return self._create_failed_result(dev_id, "用户名缺失")

        if not device_password:
            logger.error(f"{dev_id} 缺少密码！（数据库 password 字段为空）")
            return self._create_failed_result(dev_id, "密码缺失")
        # ============================================================

        device_type = dev.get("device_type") or self.config.get("default_device_type", "generic_ssh")
        model = dev.get("model", "default")
        cmd_file = dev.get("command_file")
        run_id = self.config.get("run_id")

        logger.info(f"=== 开始巡检 {dev_id} (用户名: {device_username}) ===")

        cmds = []
        if cmd_file:
            p = self.commands_dir / cmd_file
            if p.exists():
                cmds = [l.strip() for l in p.read_text(encoding="utf-8").splitlines() if l.strip()]
            else:
                logger.warning(f"{dev_id} 命令文件 {cmd_file} 不存在")
        total = len(cmds)

        if total == 0:
            logger.warning(f"{dev_id} 没有要执行的命令，跳过")
            return {"device_id": dev_id, "total": 0, "added": 0, "deleted": 0, "modified": 0}

        patrol_batch = []
        diff_batch = []
        counts = {s.value: 0 for s in State}

        conn = None
        try:
            conn = ConnectHandler(
                device_type=device_type,
                ip=ip,
                username=device_username,
                password=device_password,
                read_timeout_override=60,
                global_delay_factor=1
            )
            logger.info(f"{dev_id} SSH 连接成功")
            conn.clear_buffer(backoff=True)
        except Exception as e:
            logger.error(f"{dev_id} SSH 连接失败: {e}")
            return self._create_failed_result(dev_id, f"SSH连接失败: {e}", total, cmds)

        for cmd in cmds:
            success, output = self._ssh_execute(conn, cmd, dev_id)
            if not success:
                diff_batch.append((dev_id, cmd, State.FAILED.value, None,
                                   json.dumps([{"内容": output, "状态": State.FAILED.value}]), None, None))
                counts[State.FAILED.value] += 1
                continue

            is_struct = cmd in self.structured_cmds
            structured = None
            text_output = None
            curr_rows = None

            if is_struct:
                tmpl_name = cmd_to_template_name(cmd)
                tmpl_path = self._discover_templates(model).get(tmpl_name)
                if tmpl_path and textfsm:
                    try:
                        curr_rows = parse_with_textfsm(tmpl_path, output)
                        structured = json.dumps(curr_rows, ensure_ascii=False)
                    except Exception as e:
                        logger.warning(f"{dev_id} TextFSM 解析失败: {e}")
                        is_struct = False

            if not is_struct:
                text_output = output

            if self.force_baseline:
                state = State.ADDED.value
                sdiff = json.dumps(curr_rows, ensure_ascii=False) if curr_rows else None
                tdiff = json.dumps([{"内容": text_output, "状态": State.ADDED.value}]) if text_output else None
                baseline_text = text_output
                counts[State.ADDED.value] += len(curr_rows) if curr_rows else (output.count('\n') + 1 if output else 1)
            else:
                if cmd in self.config.get("no_compare", []):
                    if is_struct and curr_rows:
                        comp = [{**row, "状态": State.UNCHANGED.value} for row in curr_rows]
                        sdiff = json.dumps(comp, ensure_ascii=False)
                        tdiff = None
                    else:
                        comp = [{"内容": line, "状态": State.UNCHANGED.value} for line in (text_output or "").splitlines()]
                        sdiff = None
                        tdiff = json.dumps(comp, ensure_ascii=False)
                    state = State.UNCHANGED.value
                    baseline_text = None
                    for r in comp:
                        counts[r.get("状态", State.UNCHANGED.value)] += 1
                else:
                    base_struct, base_text = self.patrol_db.get_baseline_entry(dev_id, cmd)
                    if base_struct is None and base_text is None:
                        logger.error(f"{dev_id} 命令 {cmd} 无基线数据")
                        state = State.FAILED.value
                        sdiff = None
                        tdiff = json.dumps([{"内容": "无基线", "状态": State.FAILED.value}])
                        baseline_text = None
                    else:
                        if is_struct and curr_rows:
                            base_rows = json.loads(base_struct) if base_struct else []
                            key_cols = self.key_map.get(cmd, [])
                            comp = diff_structured_lcs(base_rows, curr_rows, key_cols)
                            sdiff = json.dumps(comp, ensure_ascii=False)
                            tdiff = None
                        else:
                            comp = diff_text_lines(base_text, text_output)
                            sdiff = None
                            tdiff = json.dumps(comp, ensure_ascii=False)
                        state = State.UNCHANGED.value
                        for r in comp:
                            st = r.get("状态", State.UNCHANGED.value)
                            if STATE_PRIORITY.get(st, 0) > STATE_PRIORITY.get(state, 0):
                                state = st
                            counts[st] += 1
                        baseline_text = base_text if not is_struct else None

            patrol_batch.append((dev_id, cmd, structured, text_output))
            diff_batch.append((dev_id, cmd, state, sdiff, tdiff, text_output, baseline_text))

        try:
            self.patrol_db.save_patrol_results(run_id, patrol_batch)
            self.patrol_db.save_diff_results(run_id, diff_batch)
            if self.force_baseline:
                self.patrol_db.save_baseline_results(patrol_batch)
        except Exception as e:
            logger.error(f"{dev_id} 保存数据库失败: {e}")

        added = counts.get(State.ADDED.value, 0)
        deleted = counts.get(State.DELETED.value, 0)
        modified = counts.get(State.MODIFIED.value, 0)
        self.patrol_db.save_stats(run_id, dev_id, total, added, deleted, modified)

        try:
            conn.disconnect()
        except Exception:
            pass

        logger.info(f"=== {dev_id} 完成 总:{total} 新增:{added} 缺失:{deleted} 修改:{modified} 耗时:{time.time()-start:.2f}s ===")
        return {"device_id": dev_id, "total": total, "added": added, "deleted": deleted, "modified": modified}

# -----------------------
# Excel 报告
# -----------------------
class ReportWriter:
    def __init__(self, db: PatrolDBManager, out_file: Path, config: Dict[str, Any]):
        self.db = db
        self.out_file = out_file
        self.config = config

    def write_report(self, run_id: Optional[str] = None):
        data = self.db.load_diff_for_report(run_id)
        mapping: Dict[str, str] = {}
        for sheet, cmds in self.config.get("sheet_name", {}).items():
            for cmd in cmds:
                mapping[cmd.strip()] = sheet

        new_data = defaultdict(list)
        for old_sheet, items in data.items():
            if items:
                command = items[0]["command"]
                actual_sheet_name = mapping.get(command, command)
                actual_sheet = safe_sheet_name(actual_sheet_name)
                new_data[actual_sheet].extend(items)

        wb = Workbook()
        wb.remove(wb.active)

        # 摘要页
        ws = wb.create_sheet("摘要", 0)
        ws.cell(1, 1, "设备ID").font = Font(bold=True)
        ws.cell(1, 2, "命令").font = Font(bold=True)
        ws.cell(1, 3, "新增").font = Font(bold=True)
        ws.cell(1, 4, "缺失").font = Font(bold=True)
        ws.cell(1, 5, "修改").font = Font(bold=True)
        row = 2
        for stat in self.db.get_command_stats(run_id):
            ws.cell(row, 1, stat["device_id"])
            ws.cell(row, 2, stat["command"])
            ws.cell(row, 3, stat["added"])
            ws.cell(row, 4, stat["deleted"])
            ws.cell(row, 5, stat["modified"])
            row += 1

        row += 1
        ws.cell(row, 1, "SSH 登录失败设备").font = Font(bold=True)
        row += 1
        ws.cell(row, 1, "设备ID").font = Font(bold=True)
        for dev in self.db.get_failed_devices(run_id):
            row += 1
            ws.cell(row, 1, dev)

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col if c.value is not None)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        # 命令页
        for sheet_name, items in new_data.items():
            ws = wb.create_sheet(sheet_name[:31])
            cols_set = set()
            for it in items:
                if it["structured_diff"]:
                    for r in it["structured_diff"]:
                        cols_set.update(r.keys())
                else:
                    cols_set.update(["内容", "基线输出"])
            base_order = []
            if items and items[0]["structured_diff"]:
                base_order = [k for k in items[0]["structured_diff"][0].keys() if k != "状态"]
            other = sorted(c for c in cols_set if c not in base_order and c != "状态")
            cols = ["设备ID", "命令", "状态"] + base_order + other

            for ci, c in enumerate(cols, 1):
                cell = ws.cell(1, ci, c)
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
                cell.alignment = ALIGN_CENTER

            r = 2
            for it in items:
                dev_id = it["device_id"]
                cmd = it["command"]
                if it["structured_diff"]:
                    for rec in it["structured_diff"]:
                        st = rec.get("状态", State.UNCHANGED.value)
                        vals = [dev_id, cmd, st] + [sanitize_for_excel(rec.get(c, "")) for c in base_order + other]
                        for ci, v in enumerate(vals, 1):
                            cell = ws.cell(r, ci, v)
                            if st in (State.ADDED.value, State.MODIFIED.value):
                                cell.fill = RED_FILL
                            elif st == State.DELETED.value:
                                cell.fill = GRAY_FILL
                            cell.border = THIN_BORDER
                        r += 1
                else:
                    tdiff = it["text_diff"] or []
                    st = max((x.get("状态", State.UNCHANGED.value) for x in tdiff), default=State.UNCHANGED.value,
                             key=lambda s: STATE_PRIORITY.get(s, 0))
                    vals = [dev_id, cmd, st] + [""] * len(base_order) + [
                        sanitize_for_excel(it["text_output"] or ""),
                        sanitize_for_excel(it["baseline_text"] or "")
                    ]
                    for ci, v in enumerate(vals, 1):
                        cell = ws.cell(r, ci, v)
                        if st in (State.ADDED.value, State.MODIFIED.value):
                            cell.fill = RED_FILL
                        elif st == State.DELETED.value:
                            cell.fill = GRAY_FILL
                        cell.border = THIN_BORDER
                    r += 1

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col if c.value is not None)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
            ws.freeze_panes = "A2"

        wb.save(self.out_file)
        logger.info(f"Excel 报告已生成: {self.out_file}")

# -----------------------
# 主程序
# -----------------------
def main():
    start = time.time()
    parser = argparse.ArgumentParser(description="网络设备巡检工具 - 严格使用数据库独立用户名密码")
    parser.add_argument("--group", "-g", required=True, help="设备分组名称")
    parser.add_argument("--config", "-c", default="config.yaml", help="配置文件")
    parser.add_argument("--db", default=str(DEFAULT_DB), help="巡检结果 DB")
    parser.add_argument("--workers", "-w", type=int, default=8, help="并发线程")
    parser.add_argument("--out", "-o", help="输出 Excel")
    parser.add_argument("--output-dir", default="outputs", help="输出目录")
    parser.add_argument("--baseline", "-b", action="store_true", help="强制保存基线")
    parser.add_argument("--log-level", "-l", default="ERROR", choices=["DEBUG", "INFO", "ERROR"], help="日志级别")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.getLevelName(args.log_level),
        format='%(asctime)s - %(levelname)s - [%(name)s] - %(message)s',
        handlers=[
            logging.FileHandler(LOG_DIR / "patrol.log", encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    global logger
    logger = logging.getLogger("patrol")

    cfg = load_config(Path(args.config))
    cfg.setdefault("max_retries", 1)
    cfg.setdefault("retry_delay", 1.0)
    cfg.setdefault("commands_path", "commands")
    cfg.setdefault("templates_path", "templates")
    cfg.setdefault("structured_commands", [])
    cfg.setdefault("primary_keys", {})
    cfg.setdefault("no_compare", [])
    cfg.setdefault("default_username", "admin")
    cfg.setdefault("default_device_type", "generic_ssh")
    cfg.setdefault("expect_patterns", {"default": r"[>#\$%]"})

    run_id = datetime.now(BEIJING_TZ).strftime("%Y%m%d%H%M%S")
    cfg["run_id"] = run_id

    logger.info(f"巡检任务启动 - 分组: {args.group} | 严格使用数据库独立凭证")

    patrol_db = PatrolDBManager(Path(args.db))
    device_db = DeviceDBManager(DEVICES_DB)

    devices = device_db.get_devices_by_group(args.group)
    if not devices:
        logger.error(f"分组 {args.group} 无设备")
        return

    templates_dir = Path(cfg.get("templates_path", "templates"))
    commands_dir = Path(cfg.get("commands_path", "commands"))

    patroller = DevicePatroller(patrol_db, device_db, templates_dir, commands_dir, cfg, args.baseline)

    results = []
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futures = {ex.submit(patroller.patrol_single_device, d): d for d in devices}
        for f in as_completed(futures):
            try:
                results.append(f.result())
            except Exception as e:
                logger.error(f"巡检任务异常: {e}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    if args.out is None:
        timestamp = datetime.now(BEIJING_TZ).strftime('%Y-%m-%d_%H-%M-%S')
        args.out = str(output_dir / f"{run_id}_{timestamp}_{args.group}.xlsx")

    ReportWriter(patrol_db, Path(args.out), cfg).write_report(run_id)

    patrol_db.close()
    logger.info(f"全部巡检完成，总耗时 {time.time() - start:.2f}s")


if __name__ == "__main__":
    main()
